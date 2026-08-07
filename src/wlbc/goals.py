"""The weight and step plan: targets, and progress against them.

The plan originates in a spreadsheet, but the report should not depend on a
file in Downloads staying put. `Plan.from_xlsx()` imports it once; `save()`
writes a stable `goals.json` beside the project that everything else reads.
"""

from __future__ import annotations

import datetime as dt
import json
import math
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

DEFAULT_GOALS_PATH = Path("goals.json")

KG_PER_LB = 0.45359237


class GoalsError(Exception):
    """The plan is missing or unreadable."""


@dataclass
class StepPlan:
    """A step target that ramps on a fixed schedule and then holds."""

    baseline: int = 3000
    goal: int = 12000
    increment: int = 1000
    every_weeks: int = 2

    def target(self, week: int) -> int:
        """Steps for a given plan week.

        Week 0 sits at baseline; the first bump lands in week 1, then every
        `every_weeks` after that. Ceiling division reproduces that off-by-one
        opening exactly (weeks 1-2 share a target, 3-4 share the next).
        """
        if week < 0:
            week = 0
        if self.every_weeks < 1:
            raise ValueError("every_weeks must be at least 1")
        steps = self.baseline + self.increment * math.ceil(week / self.every_weeks)
        return min(self.goal, steps)


@dataclass
class Plan:
    """A linear weight target over a fixed horizon, plus a step ramp."""

    start_date: dt.date
    start_weight: float
    goal_weight: float
    plan_weeks: int = 52
    units: str = "lb"
    steps: StepPlan = field(default_factory=StepPlan)

    # Context carried through from the spreadsheet, shown but not computed on.
    height_in: float | None = None
    bmr: float | None = None
    activity_factor: float | None = None
    maintenance_calories: float | None = None
    target_calories: float | None = None

    # -- derived ---------------------------------------------------------

    @property
    def end_date(self) -> dt.date:
        return self.start_date + dt.timedelta(weeks=self.plan_weeks)

    @property
    def total_change(self) -> float:
        return self.goal_weight - self.start_weight

    @property
    def per_week(self) -> float:
        return self.total_change / self.plan_weeks if self.plan_weeks else 0.0

    def week_of(self, day: dt.date) -> float:
        return (day - self.start_date).days / 7.0

    def target_weight(self, day: dt.date) -> float | None:
        """Target weight on a day, or None outside the plan window.

        Linear between start and goal. Past the end date the target holds at
        the goal rather than continuing down — the plan stops, the line doesn't
        keep falling.
        """
        if day < self.start_date:
            return None
        week = min(self.week_of(day), float(self.plan_weeks))
        return self.start_weight + self.total_change * (week / self.plan_weeks)

    def target_steps(self, day: dt.date) -> int | None:
        if day < self.start_date:
            return None
        return self.steps.target(int(self.week_of(day)))

    def in_units(self, units: str) -> "Plan":
        """Return the plan converted to kg or lb."""
        if units == self.units:
            return self
        if units not in ("kg", "lb"):
            raise ValueError(f"units must be 'kg' or 'lb'; got {units!r}")
        factor = KG_PER_LB if units == "kg" else 1 / KG_PER_LB
        clone = Plan(**{**asdict(self), "steps": self.steps, "start_date": self.start_date})
        clone.start_weight *= factor
        clone.goal_weight *= factor
        clone.units = units
        return clone

    # -- persistence -----------------------------------------------------

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["start_date"] = self.start_date.isoformat()
        return data

    def save(self, path: Path | str = DEFAULT_GOALS_PATH) -> Path:
        path = Path(path)
        path.write_text(json.dumps(self.to_dict(), indent=2))
        return path

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "Plan":
        data = dict(data)
        data["start_date"] = dt.date.fromisoformat(str(data["start_date"])[:10])
        data["steps"] = StepPlan(**(data.get("steps") or {}))
        known = {f for f in cls.__dataclass_fields__}
        return cls(**{k: v for k, v in data.items() if k in known})

    @classmethod
    def load(cls, path: Path | str = DEFAULT_GOALS_PATH) -> "Plan":
        path = Path(path)
        if not path.is_file():
            raise GoalsError(
                f"No plan at {path}. Import one with `wlbc goals import <file.xlsx>`."
            )
        try:
            return cls.from_dict(json.loads(path.read_text()))
        except (json.JSONDecodeError, KeyError, TypeError, ValueError) as exc:
            raise GoalsError(f"Plan at {path} is unreadable: {exc}") from exc

    # -- import ----------------------------------------------------------

    @classmethod
    def from_xlsx(cls, path: Path | str, sheet: str = "Dashboard") -> "Plan":
        """Read the plan out of the tracker spreadsheet's Dashboard sheet."""
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - depends on the extra
            raise GoalsError(
                "Reading .xlsx needs openpyxl. Install it with: pip install 'wlbc[goals]'"
            ) from exc

        path = Path(path)
        if not path.is_file():
            raise GoalsError(f"No such file: {path}")

        workbook = openpyxl.load_workbook(path, data_only=True)
        if sheet not in workbook.sheetnames:
            raise GoalsError(
                f"{path.name} has no {sheet!r} sheet. Found: {', '.join(workbook.sheetnames)}"
            )

        # The Dashboard is label/value pairs down two columns.
        values: dict[str, Any] = {}
        for row in workbook[sheet].iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            if len(cells) >= 2 and isinstance(cells[0], str):
                values.setdefault(cells[0].strip().lower(), cells[1])

        def need(label: str) -> Any:
            for key, value in values.items():
                if key.startswith(label.lower()):
                    return value
            raise GoalsError(f"{path.name} is missing a {label!r} row on the {sheet} sheet.")

        def maybe(label: str) -> Any:
            try:
                return need(label)
            except GoalsError:
                return None

        start = need("start date")
        start_date = start.date() if isinstance(start, dt.datetime) else dt.date.fromisoformat(str(start)[:10])

        return cls(
            start_date=start_date,
            start_weight=float(need("start weight")),
            goal_weight=float(need("goal weight")),
            plan_weeks=int(need("plan length")),
            units="lb" if "(lb)" in next(k for k in values if k.startswith("start weight")) else "kg",
            steps=StepPlan(
                baseline=int(need("baseline daily steps")),
                goal=int(need("goal daily steps")),
                increment=int(need("step increase increment")),
                every_weeks=int(need("step increase frequency")),
            ),
            height_in=_float_or_none(maybe("height")),
            bmr=_float_or_none(maybe("measured bmr")),
            activity_factor=_float_or_none(maybe("activity factor")),
            maintenance_calories=_float_or_none(maybe("maintenance calories")),
            target_calories=_float_or_none(maybe("recommended target calories")),
        )


def _float_or_none(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


@dataclass
class Progress:
    """Where actual weight sits against the plan."""

    day: dt.date | None = None
    week: float = 0.0
    actual: float | None = None
    target: float | None = None
    delta: float | None = None          # actual - target; negative is ahead
    lost_so_far: float | None = None
    remaining: float | None = None
    actual_per_week: float | None = None
    required_per_week: float | None = None
    projected_end_weight: float | None = None
    on_track: bool | None = None

    @property
    def status(self) -> str:
        if self.delta is None:
            return "no data"
        if abs(self.delta) < 0.25:
            return "on track"
        return "ahead of plan" if self.delta < 0 else "behind plan"


def evaluate(
    plan: Plan,
    records,
    actual_per_week: float | None = None,
) -> Progress:
    """Compare the latest weigh-in against the plan.

    `actual_per_week` comes from the least-squares fit over real weigh-ins;
    with fewer than two it is None and the projection is skipped rather than
    extrapolated from a single point.
    """
    weighed = [r for r in records if getattr(r, "weight_kg", None) is not None]
    if not weighed:
        return Progress()

    latest = weighed[-1]
    # Prefer the smoothed value when there is one; a lone reading is raw.
    actual = latest.trend.get("weight_kg") or latest.weight_kg
    target = plan.target_weight(latest.day)

    progress = Progress(
        day=latest.day,
        week=plan.week_of(latest.day),
        actual=actual,
        target=target,
        lost_so_far=actual - plan.start_weight if actual is not None else None,
        remaining=plan.goal_weight - actual if actual is not None else None,
        required_per_week=plan.per_week,
        actual_per_week=actual_per_week,
    )
    if actual is not None and target is not None:
        progress.delta = actual - target
        progress.on_track = progress.delta <= 0.25

    if actual is not None and actual_per_week is not None:
        weeks_left = max(0.0, plan.plan_weeks - progress.week)
        progress.projected_end_weight = actual + actual_per_week * weeks_left

    return progress
